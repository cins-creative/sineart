#!/usr/bin/env python3
"""
Đọc file Excel «BCTC 2025.xlsx» sheet KQKD 2025, map vào cột tc_bao_cao_tai_chinh.

Giả định thiếu chi tiết TTM/HH/BCM Online–Offline: chia đều 3 môn,
mỗi môn áp tỉ lệ 70% Online / 30% Offline trên «tổng môn» P,
sau đó chỉnh làm tròn để:
  dtTTMOnline+dtHHOnline+dtBCMOnline + dtKids (+ dtBackground) == Doanh thu Online (Excel)
  dtTTMOffline+dtHHOffline+dtBCMOffline == Doanh thu Offline (Excel)

dtBackground = 0 (không có dòng riêng trong file mẫu).

Chi phí / lương map theo nhãn hàng Excel (xem MAP_ROWS).

Chạy: python scripts/generate-bctc-2025-supabase-sql.py [đường_dẫn_xlsx]
Xuất: scripts/output/bctc_2025_seed.sql
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Cần: pip install openpyxl")
    sys.exit(1)

MONTH_HEADERS = [f"Tháng {i}" for i in range(1, 13)]
MON_COL_START = 6  # cột Tháng 1 trong sheet đã đọc (0-based)


def nz(x) -> float:
    if x is None:
        return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def norm_label(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def find_row(ws, predicate):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cell0 = row[0] if row else None
        if predicate(cell0):
            return i, row
    return None, None


def split_three_mon_70_30(online_total: float, offline_total: float, kids: float, bg: float):
    """
    online_total = Excel 4.1 (đã gồm kids online trong báo cáo KT hay không?)
    Giả định: 4.1 = TTM_o+HH_o+BCM_o+Kids+BG (đúng với định nghĩa app).
    offline_total = Excel 4.2 = TTM_f+HH_f+BCM_f

    P = (TTM_o+TTM_f) = (HH_o+HH_f) = (BCM_o+BCM_f) sau khi chia đều 3 môn.
    """
    so = online_total - kids - bg  # chỉ 3 môn online
    sf = offline_total

    # Tổng tiền 3 môn (mỗi môn có TTM_o+TTM_f = P)
    combined_three = so + sf
    if combined_three <= 0:
        return {
            "dtTTMOnline": 0,
            "dtTTMOffline": 0,
            "dtHHOnline": 0,
            "dtHHOffline": 0,
            "dtBCMOnline": 0,
            "dtBCMOffline": 0,
        }

    P = combined_three / 3.0
    keys = [
        ("dtTTMOnline", "dtTTMOffline"),
        ("dtHHOnline", "dtHHOffline"),
        ("dtBCMOnline", "dtBCMOffline"),
    ]
    out = {}
    sum_o = 0.0
    sum_f = 0.0
    for ko, kf in keys:
        on = round(P * 0.7)
        off = round(P * 0.3)
        out[ko] = on
        out[kf] = off
        sum_o += on
        sum_f += off

    # Hiệu chỉnh để khớp Excel (ưu tiên khớp online 3 môn trước)
    d_o = int(round(so - sum_o))
    d_f = int(round(sf - sum_f))

    # Phân phối chênh lệch vào các ô theo thứ tự (TTM, HH, BCM)
    adjust_order_o = ["dtTTMOnline", "dtHHOnline", "dtBCMOnline"]
    adjust_order_f = ["dtTTMOffline", "dtHHOffline", "dtBCMOffline"]

    def spread(diff: int, order):
        i = 0
        while diff != 0 and i < 100:
            k = order[i % len(order)]
            step = 1 if diff > 0 else -1
            if out[k] + step >= 0:
                out[k] += step
                diff -= step
            i += 1

    spread(d_o, adjust_order_o)
    spread(d_f, adjust_order_f)

    return out


def row_get_months(row, n_months=12):
    out = []
    for m in range(n_months):
        idx = MON_COL_START + m
        out.append(nz(row[idx] if idx < len(row) else None))
    return out


def sql_escape(s: str) -> str:
    return s.replace("'", "''")


def main():
    # Tránh UnicodeEncodeError khi print trên Windows (cp1252)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    xlsx = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else r"c:\Users\DELL\Downloads\BCTC 2025.xlsx"
    )
    out_dir = Path(__file__).resolve().parent / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_sql = out_dir / "bctc_2025_seed.sql"

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "KQKD 2025" not in wb.sheetnames:
        print("Không thấy sheet 'KQKD 2025'. Có:", wb.sheetnames)
        sys.exit(1)
    ws = wb["KQKD 2025"]

    # --- Thu thập label -> list 12 tháng ---
    def by_prefix(prefix: str):
        def pred(cell):
            return norm_label(cell).startswith(norm_label(prefix))

        _, r = find_row(ws, pred)
        return row_get_months(r) if r else [0.0] * 12

    # Doanh thu (theo dump Excel trước đó)
    rev_gross = by_prefix("1.")  # không dùng trực tiếp nếu đã có 4.x
    ck = by_prefix("2.")  # chiết khấu
    hang_tra = by_prefix("3.")  # hàng trả lại
    dt_online_excel = by_prefix("4.1")
    dt_offline_excel = by_prefix("4.2")
    dt_kids_excel = by_prefix("4.3")
    dt_hoa_cu_excel = by_prefix("4.4")  # doanh thu bán họa cụ
    dt_tc_excel = by_prefix("7.")  # hoạt động tài chính

    # Chi phí — nhãn có thể khác dấu chấm
    cp_luong_gv = by_prefix("lương giảng viên")
    cp_luong_vh = by_prefix("lương vận hành")
    cp_trich_truoc = by_prefix("chi phí trích trước")
    cp_qua = by_prefix("chi phí quà")
    cp_dao_tao = by_prefix("chi phí dụng cụ")
    cp_mkt_web = by_prefix("chi phí mkt")
    cp_video = by_prefix("chi phí video")
    cp_khau_hao = by_prefix("chi phí khấu hao")
    cp_dien = by_prefix("chi phí điện")
    cp_mat_bang = by_prefix("chi phí mặt bằng")
    cp_ngan_hang = by_prefix("chi phí ngân hàng")
    cp_kids_hc = by_prefix("chi phí họa cụ dùng cho lớp kids")
    cp_nhan_hang = by_prefix("chi phí nhận hàng")
    cp_tien_khac = by_prefix("chi phí bằng tiền khác")

    lines = []
    lines.append("-- Seed tc_bao_cao_tai_chinh — BCTC 2025 (sinh tự động)")
    lines.append("-- Xóa trước INSERT để tránh trùng UNIQUE (nam, thang). Backup nếu cần.")
    lines.append("BEGIN;")
    lines.append("DELETE FROM tc_bao_cao_tai_chinh WHERE nam = '2025';")
    lines.append("")

    THANG_VI = [f"Tháng {i}" for i in range(1, 13)]

    for mi in range(12):
        thang = THANG_VI[mi]
        on = dt_online_excel[mi]
        off = dt_offline_excel[mi]
        kids = dt_kids_excel[mi]
        bg = 0.0

        dt_split = split_three_mon_70_30(on, off, kids, bg)

        row_db = {
            "nam": "2025",
            "thang": thang,
            # DT lớp (đã phân 70/30)
            "dt_ttm_onl": dt_split["dtTTMOnline"],
            "dt_ttm_off": dt_split["dtTTMOffline"],
            "dt_hh_onl": dt_split["dtHHOnline"],
            "dt_hh_off": dt_split["dtHHOffline"],
            "dt_bcm_onl": dt_split["dtBCMOnline"],
            "dt_bcm_off": dt_split["dtBCMOffline"],
            "dt_kids": int(round(kids)),
            "dt_background": 0,
            "dt_dich_vu": 0,
            "dt_marketplace": int(round(dt_hoa_cu_excel[mi])),  # họa cụ bán → marketplace (tách khỏi dt_hh_off)
            "dt_hoat_dong_tc": int(round(dt_tc_excel[mi])),
            "chietkhau_khuyenmai": int(round(ck[mi])),
            "hangban_tralai": int(round(hang_tra[mi])),
            # Lương: GV / VH
            "luong_gv_luyenthi": int(round(cp_luong_gv[mi])),
            "luong_gv_background": 0,
            "luong_vh_luyenthi": int(round(cp_luong_vh[mi] * 0.5)),
            "luong_vh_web": int(round(cp_luong_vh[mi] * 0.5)),
            "bhxh_nhanvien": 0,
            "thue_vat_hocvien": 0,
            "thue_cungcapdichvu": 0,
            "cp_trich_truoc": int(round(cp_trich_truoc[mi])),
            "cp_qua_sinhnhat": int(round(cp_qua[mi])),
            "cp_daotao": int(round(cp_dao_tao[mi] + cp_video[mi] + cp_kids_hc[mi])),
            "cp_website": int(round(cp_mkt_web[mi] / 3)),
            "cp_phanmem": int(round(cp_mkt_web[mi] / 3)),
            "cp_khac": int(round(cp_mkt_web[mi] / 3)),  # marketing + phần còn map vào cp_khac (cp_marketing trong app)
            "cp_khauhao_tscd": int(round(cp_khau_hao[mi])),
            "cp_diennuoc_internet": int(round(cp_dien[mi])),
            "cp_matbang": int(round(cp_mat_bang[mi])),
            "cp_nganhang": int(round(cp_ngan_hang[mi])),
            "cp_khac_dup": int(round(cp_tien_khac[mi] + cp_nhan_hang[mi])),  # will fix key below
        }
        # Gộp phần cp_khac — app chỉ có một cp_khac cho marketing vs tien khác; đơn giản hóa:
        mkt_total = int(round(cp_mkt_web[mi]))
        row_db["cp_khac"] = mkt_total  # toàn bộ MKT/Web/PM vào cp_khac (cột cp_khac trong DB)
        row_db.pop("cp_khac_dup", None)
        row_db["cp_website"] = 0
        row_db["cp_phanmem"] = 0

        # Postgres một cột cp_khac; gộp MKT/Web/PM + tiền khác + nhận hàng (xem final bên dưới)
        final = {
            "nam": row_db["nam"],
            "thang": row_db["thang"],
            "dt_ttm_onl": row_db["dt_ttm_onl"],
            "dt_ttm_off": row_db["dt_ttm_off"],
            "dt_hh_onl": row_db["dt_hh_onl"],
            "dt_hh_off": row_db["dt_hh_off"],
            "dt_bcm_onl": row_db["dt_bcm_onl"],
            "dt_bcm_off": row_db["dt_bcm_off"],
            "dt_kids": row_db["dt_kids"],
            "dt_background": row_db["dt_background"],
            "dt_dich_vu": row_db["dt_dich_vu"],
            "dt_marketplace": row_db["dt_marketplace"],
            "dt_hoat_dong_tc": row_db["dt_hoat_dong_tc"],
            "chietkhau_khuyenmai": row_db["chietkhau_khuyenmai"],
            "hangban_tralai": row_db["hangban_tralai"],
            "luong_gv_luyenthi": row_db["luong_gv_luyenthi"],
            "luong_gv_background": row_db["luong_gv_background"],
            "luong_vh_luyenthi": row_db["luong_vh_luyenthi"],
            "luong_vh_web": row_db["luong_vh_web"],
            "bhxh_nhanvien": row_db["bhxh_nhanvien"],
            "thue_vat_hocvien": row_db["thue_vat_hocvien"],
            "thue_cungcapdichvu": row_db["thue_cungcapdichvu"],
            "cp_trich_truoc": row_db["cp_trich_truoc"],
            "cp_qua_sinhnhat": row_db["cp_qua_sinhnhat"],
            "cp_daotao": row_db["cp_daotao"],
            "cp_website": 0,
            "cp_phanmem": 0,
            "cp_khac": int(round(cp_mkt_web[mi])) + int(round(cp_tien_khac[mi])) + int(round(cp_nhan_hang[mi])),
            "cp_khauhao_tscd": row_db["cp_khauhao_tscd"],
            "cp_diennuoc_internet": row_db["cp_diennuoc_internet"],
            "cp_matbang": row_db["cp_matbang"],
            "cp_nganhang": row_db["cp_nganhang"],
        }

        col_list = ", ".join(final.keys())
        val_list = ", ".join(
            str(int(v)) if isinstance(v, (int, float)) and not isinstance(v, bool) else f"'{sql_escape(str(v))}'"
            for v in final.values()
        )
        lines.append(
            f"INSERT INTO tc_bao_cao_tai_chinh ({col_list}) VALUES ({val_list});"
        )

    lines.append("COMMIT;")
    out_sql.write_text("\n".join(lines), encoding="utf-8")
    print("Đã ghi:", out_sql)

    # Phân tích ngắn
    meta = {
        "source": str(xlsx),
        "sheet": "KQKD 2025",
        "rules": {
            "three_subjects_equal_P": "P = (Sum Online 3 môn + Sum Offline 3 môn) / 3 với Online3 = Excel4.1 - Kids - BG, Offline3 = Excel4.2",
            "per_subject": "Online=round(0.7*P), Offline=round(0.3*P), then adjust để khớp tổng Excel",
            "dt_marketplace": "dòng 4.4 Doanh thu bán họa cụ",
            "dt_hh_off_conflict": "App map dtHoaCu -> dt_hh_off; script dùng dt_marketplace cho họa cụ để tránh ghi đè HH Offline",
        },
    }
    (out_dir / "bctc_2025_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
